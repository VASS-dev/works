"""Печатные формы — Excel-документы для руководства и рабочих.

Каждая форма форматирована для печати на A4 с рамкой, заголовками,
блоками подписей и настройками страницы.
"""
from __future__ import annotations

from collections import OrderedDict
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, selectinload

from app.models import Contract, ExecutionEntry, OrderLine, PlanEntry
from app.services.balances import balance_bulk
from app.services.pricing import unit_price
from app.services.statuses import ExecutionStatus


# ── Стили ────────────────────────────────────────────────────────────────────

_T = Side(style="thin")
_M = Side(style="medium")
_BORDER = Border(left=_T, right=_T, top=_T, bottom=_T)
_BOTTOM_THICK = Border(left=_T, right=_T, top=_T, bottom=_M)
_ONLY_BOTTOM = Border(bottom=_M)

_FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")   # синеватый — шапка таблицы
_FILL_GROUP  = PatternFill("solid", fgColor="EEF2F7")   # серый — строка группы
_FILL_TOTAL  = PatternFill("solid", fgColor="E2E8F0")   # итоговая строка
_FILL_SIGNED = PatternFill("solid", fgColor="E8F5E9")   # подписано — зелёный

_ARIAL = "Arial"

MONTHS_RU = [
    "", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
]
MONTHS_GEN = [
    "", "января", "февраля", "марта", "апреля", "мая", "июня",
    "июля", "августа", "сентября", "октября", "ноября", "декабря",
]


def _period_ru(period: str, genitive: bool = False) -> str:
    y, m = period.split("-")
    w = MONTHS_GEN if genitive else MONTHS_RU
    return f"{w[int(m)]} {y} г."


# ── Helpers ───────────────────────────────────────────────────────────────────

def _cell(ws, row: int, col: int, value=None, *,
          bold=False, size=9, h_align="left", wrap=False,
          fill=None, border=True, num_fmt=None, italic=False) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size, name=_ARIAL, italic=italic)
    c.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=wrap)
    if border:
        c.border = _BORDER
    if fill:
        c.fill = fill
    if num_fmt:
        c.number_format = num_fmt


def _merge(ws, r1, c1, r2, c2, value=None, *, bold=False, size=11,
           h_align="center", fill=None, border=False) -> None:
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    c.font = Font(bold=bold, size=size, name=_ARIAL)
    c.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=True)
    if fill:
        c.fill = fill
    if border:
        c.border = _BORDER


def _row_h(ws, row: int, h: float) -> None:
    ws.row_dimensions[row].height = h


def _col_w(ws, col: int, w: float) -> None:
    ws.column_dimensions[get_column_letter(col)].width = w


def _table_borders(ws, r1: int, r2: int, c1: int, c2: int) -> None:
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for cell in row:
            cell.border = _BORDER


def _a4(ws, landscape: bool = False) -> None:
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.6
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.6
    ws.page_margins.bottom = 0.6
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2


def _finish(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# ФОРМА 1: Задание на производство работ
# ─────────────────────────────────────────────────────────────────────────────

def _build_work_order_sheet(wb: Workbook, ws, session: Session,
                             contract: Contract, period: str) -> None:
    """Заполняет лист ws данными задания на производство работ для одного договора."""
    N = 9
    widths = [4, 42, 7, 9, 9, 9, 10, 13, 13]
    for i, w in enumerate(widths, 1):
        _col_w(ws, i, w)

    ols = (
        session.query(OrderLine)
        .filter(OrderLine.contract_id == contract.id)
        .options(selectinload(OrderLine.work_type))
        .order_by(OrderLine.id)
        .all()
    )
    plan_map = {
        pe.order_line_id: pe
        for pe in session.query(PlanEntry).filter(
            PlanEntry.order_line_id.in_([ol.id for ol in ols]),
            PlanEntry.period == period,
        ).all()
    }
    balances = balance_bulk(session, [ol.id for ol in ols])

    # ── Шапка документа ─────────────────────────────────────────────
    _row_h(ws, 1, 4)
    _row_h(ws, 2, 28)
    _merge(ws, 2, 1, 2, N, "ЗАДАНИЕ НА ПРОИЗВОДСТВО РАБОТ", bold=True, size=14)
    _row_h(ws, 3, 16)
    _merge(ws, 3, 1, 3, N, _period_ru(period).upper(), bold=False, size=11)
    _row_h(ws, 4, 6)

    def meta(row, label, value):
        _row_h(ws, row, 15)
        _merge(ws, row, 1, row, 2, label, bold=True, size=10, h_align="left")
        _merge(ws, row, 3, row, N, value or "—", bold=False, size=10, h_align="left")

    contract_ref = contract.contract_no or contract.number
    if contract.contract_date:
        contract_ref += f" от {contract.contract_date.strftime('%d.%m.%Y')}"

    meta(5, "Подрядчик:", contract.counterparty)
    meta(6, "Объект:", contract.project)
    meta(7, "Договор №:", contract_ref)
    _row_h(ws, 8, 6)

    # ── Заголовок таблицы ────────────────────────────────────────────
    headers = [
        "№", "Наименование работы", "Ед.\nизм.",
        "Кол-во\nпо дог.", "Выпол-\nнено", "Остаток",
        "На период\n(кол-во)", "Ед. цена,\nруб.", "Сумма,\nруб.",
    ]
    r = 9
    _row_h(ws, r, 30)
    for i, h in enumerate(headers, 1):
        _cell(ws, r, i, h, bold=True, size=9, h_align="center", wrap=True, fill=_FILL_HEADER)

    # ── Строки данных ────────────────────────────────────────────────
    r = 10
    total_qty = 0.0
    total_sum = 0.0

    for n, ol in enumerate(ols, 1):
        b = balances.get(ol.id)
        pe = plan_map.get(ol.id)
        u = unit_price(ol)
        plan_qty = float(pe.plan_qty) if pe and pe.plan_qty is not None else None
        plan_sum = round(plan_qty * u, 2) if plan_qty is not None and u else None
        qty_remaining = float(b.qty_remaining) if b and b.qty_remaining is not None else None

        if plan_qty:
            total_qty += plan_qty
        if plan_sum:
            total_sum += plan_sum

        unit_name = (ol.work_type.unit if ol.work_type and ol.work_type.unit else "—")

        _cell(ws, r, 1, n, h_align="center")
        _cell(ws, r, 2, ol.description, h_align="left", wrap=True)
        _cell(ws, r, 3, unit_name, h_align="center")
        _cell(ws, r, 4, float(ol.qty_ordered or 0), h_align="right", num_fmt="#,##0.00")
        _cell(ws, r, 5, float(b.qty_done) if b else 0.0, h_align="right", num_fmt="#,##0.00")
        _cell(ws, r, 6, qty_remaining, h_align="right", num_fmt="#,##0.00")
        _cell(ws, r, 7, plan_qty, h_align="right", num_fmt="#,##0.00")
        _cell(ws, r, 8, round(u, 2) if u else None, h_align="right", num_fmt="#,##0.00")
        _cell(ws, r, 9, plan_sum, h_align="right", num_fmt="#,##0.00")
        r += 1

    # ── Итого ────────────────────────────────────────────────────────
    _row_h(ws, r, 16)
    _merge(ws, r, 1, r, 6, "ИТОГО", bold=True, size=10, fill=_FILL_TOTAL, border=True)
    _table_borders(ws, r, r, 1, 6)
    _cell(ws, r, 7, round(total_qty, 2), bold=True, h_align="right", fill=_FILL_TOTAL, num_fmt="#,##0.00")
    _cell(ws, r, 8, None, fill=_FILL_TOTAL)
    _cell(ws, r, 9, round(total_sum, 2), bold=True, h_align="right", fill=_FILL_TOTAL, num_fmt="#,##0.00")
    r += 2

    # ── Блок подписей ─────────────────────────────────────────────────
    def sig_line(row, label):
        _row_h(ws, row, 22)
        _merge(ws, row, 1, row, 3, label, bold=False, size=10, h_align="left")
        for col, text in [(4, "должность"), (6, "ФИО"), (8, "дата")]:
            ws.cell(row=row, column=col, value=text).font = Font(size=8, name=_ARIAL, italic=True, color="808080")
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="bottom")
            ws.cell(row=row, column=col).border = Border(bottom=_T)
        ws.column_dimensions[get_column_letter(5)].width = 1
        ws.cell(row=row, column=5, value="/").alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=7, value="/").alignment = Alignment(horizontal="center", vertical="center")

    _row_h(ws, r, 6); r += 1
    sig_line(r, "Задание выдал:"); r += 1
    _row_h(ws, r, 8); r += 1
    sig_line(r, "Задание принял:"); r += 1
    _row_h(ws, r, 8); r += 1
    sig_line(r, "Дата выдачи задания:")

    # ── Настройки печати ──────────────────────────────────────────────
    _a4(ws, landscape=False)
    ws.print_title_rows = "9:9"
    ws.freeze_panes = "A10"


def export_work_order(session: Session, period: str,
                      contract_id: Optional[int] = None) -> bytes:
    """Задание на производство работ.

    Если contract_id указан — один лист для этого договора.
    Если None — один лист на каждый договор, у которого есть план на период.
    """
    if contract_id is not None:
        contracts = [session.get(Contract, contract_id)]
        if contracts[0] is None:
            raise ValueError(f"Договор {contract_id} не найден")
    else:
        # Найти все договоры, у которых есть PlanEntry на период
        contract_ids = (
            session.query(OrderLine.contract_id)
            .join(PlanEntry, PlanEntry.order_line_id == OrderLine.id)
            .filter(PlanEntry.period == period, PlanEntry.plan_qty.isnot(None))
            .distinct()
            .all()
        )
        cids = [row[0] for row in contract_ids]
        contracts = (
            session.query(Contract)
            .filter(Contract.id.in_(cids))
            .order_by(Contract.counterparty)
            .all()
        )

    wb = Workbook()
    first = True
    for contract in contracts:
        if first:
            ws = wb.active
            first = False
        else:
            ws = wb.create_sheet()
        # Название листа — первые 31 символа counterparty (ограничение Excel)
        sheet_name = (contract.counterparty or contract.number or str(contract.id))[:31]
        ws.title = sheet_name
        _build_work_order_sheet(wb, ws, session, contract, period)

    return _finish(wb)


# ─────────────────────────────────────────────────────────────────────────────
# ФОРМА 2: Сводный производственный план на период
# ─────────────────────────────────────────────────────────────────────────────

def export_monthly_plan_form(session: Session, period: str,
                              contract_id: Optional[int] = None) -> bytes:
    """Сводный производственный план — для руководства.

    Все плановые работы за период сгруппированы по договорам с промежуточными
    итогами и итоговой строкой. Показывает только строки с планом > 0.
    """
    qset = (
        session.query(PlanEntry)
        .filter(PlanEntry.period == period, PlanEntry.plan_qty.isnot(None))
        .options(
            selectinload(PlanEntry.order_line)
            .selectinload(OrderLine.contract),
            selectinload(PlanEntry.order_line)
            .selectinload(OrderLine.work_type),
        )
        .order_by(PlanEntry.order_line_id)
    )
    if contract_id:
        from sqlalchemy.orm import aliased
        qset = qset.join(OrderLine, PlanEntry.order_line_id == OrderLine.id).filter(
            OrderLine.contract_id == contract_id
        )
    entries = qset.all()

    # Группировка по договору

    groups: dict[int, list[PlanEntry]] = OrderedDict()
    for pe in entries:
        cid = pe.order_line.contract_id
        groups.setdefault(cid, []).append(pe)

    wb = Workbook()
    ws = wb.active
    ws.title = "Производственный план"
    N = 8

    # Ширины
    for i, w in enumerate([4, 38, 22, 7, 10, 12, 13, 13], 1):
        _col_w(ws, i, w)

    # ── Заголовок ────────────────────────────────────────────────────
    _row_h(ws, 1, 4)
    _row_h(ws, 2, 28)
    _merge(ws, 2, 1, 2, N, "СВОДНЫЙ ПРОИЗВОДСТВЕННЫЙ ПЛАН", bold=True, size=14)
    _row_h(ws, 3, 16)
    _merge(ws, 3, 1, 3, N, f"на {_period_ru(period, genitive=True).lower()}", size=11)
    _row_h(ws, 4, 6)

    # ── Шапка таблицы ────────────────────────────────────────────────
    headers = [
        "№", "Наименование работы", "Объект / Договор",
        "Ед.\nизм.", "Кол-во\nплан", "Ед. цена,\nруб.",
        "Сумма\nплан, руб.", "Статус\nплана",
    ]
    r = 5
    _row_h(ws, r, 28)
    for i, h in enumerate(headers, 1):
        _cell(ws, r, i, h, bold=True, size=9, h_align="center", wrap=True, fill=_FILL_HEADER)

    # ── Данные ───────────────────────────────────────────────────────
    r = 6
    grand_qty = 0.0
    grand_sum = 0.0
    n_global = 0

    for cid, pes in groups.items():
        first_pe = pes[0]
        c = first_pe.order_line.contract

        # Строка-разделитель договора
        _row_h(ws, r, 15)
        contract_label = f"{c.counterparty or '—'}   {c.contract_no or c.number}"
        _merge(ws, r, 1, r, N, contract_label, bold=True, size=9,
               fill=_FILL_GROUP, h_align="left", border=True)
        _table_borders(ws, r, r, 1, N)
        r += 1

        group_qty = 0.0
        group_sum = 0.0

        for pe in pes:
            n_global += 1
            ol = pe.order_line
            u = unit_price(ol)
            qty = float(pe.plan_qty)
            s = round(qty * u, 2) if u else 0.0
            group_qty += qty
            group_sum += s
            status_label = "утверждён" if pe.status == "approved" else "черновик"

            _cell(ws, r, 1, n_global, h_align="center")
            _cell(ws, r, 2, ol.description, h_align="left", wrap=True)
            _cell(ws, r, 3, c.project or "", h_align="left", wrap=True)
            _cell(ws, r, 4, ol.work_type.unit if ol.work_type and ol.work_type.unit else "—", h_align="center")
            _cell(ws, r, 5, qty, h_align="right", num_fmt="#,##0.00")
            _cell(ws, r, 6, round(u, 2) if u else None, h_align="right", num_fmt="#,##0.00")
            _cell(ws, r, 7, s if s else None, h_align="right", num_fmt="#,##0.00")
            _cell(ws, r, 8, status_label, h_align="center",
                  fill=_FILL_SIGNED if pe.status == "approved" else None)
            r += 1

        # Итого по договору
        _row_h(ws, r, 14)
        _merge(ws, r, 1, r, 4, f"Итого по: {c.counterparty or '—'}", bold=True, size=9,
               fill=_FILL_TOTAL, h_align="left", border=True)
        _table_borders(ws, r, r, 1, 4)
        _cell(ws, r, 5, round(group_qty, 2), bold=True, fill=_FILL_TOTAL, h_align="right", num_fmt="#,##0.00")
        _cell(ws, r, 6, None, fill=_FILL_TOTAL)
        _cell(ws, r, 7, round(group_sum, 2), bold=True, fill=_FILL_TOTAL, h_align="right", num_fmt="#,##0.00")
        _cell(ws, r, 8, None, fill=_FILL_TOTAL)

        grand_qty += group_qty
        grand_sum += group_sum
        r += 1

    # ── ИТОГО ────────────────────────────────────────────────────────
    _row_h(ws, r, 16)
    _merge(ws, r, 1, r, 4, "ИТОГО ПО ВСЕМ ДОГОВОРАМ", bold=True, size=10,
           fill=_FILL_HEADER, h_align="left", border=True)
    _table_borders(ws, r, r, 1, 4)
    _cell(ws, r, 5, round(grand_qty, 2), bold=True, size=10, fill=_FILL_HEADER, h_align="right", num_fmt="#,##0.00")
    _cell(ws, r, 6, None, fill=_FILL_HEADER)
    _cell(ws, r, 7, round(grand_sum, 2), bold=True, size=10, fill=_FILL_HEADER, h_align="right", num_fmt="#,##0.00")
    _cell(ws, r, 8, None, fill=_FILL_HEADER)

    _a4(ws, landscape=True)
    ws.print_title_rows = "5:5"
    ws.freeze_panes = "A6"

    return _finish(wb)


# ─────────────────────────────────────────────────────────────────────────────
# ФОРМА 3: Оперативный отчёт о выполнении работ
# ─────────────────────────────────────────────────────────────────────────────

def export_progress_report(session: Session, period: str,
                            contract_id: Optional[int] = None) -> bytes:
    """Оперативный отчёт о выполнении работ за период — для руководства.

    Сравнение план/факт по каждой работе. Подписанные акты выделены.
    Содержит итоги по договорам и общий итог.
    """
    qset = (
        session.query(ExecutionEntry)
        .join(PlanEntry, ExecutionEntry.plan_entry_id == PlanEntry.id)
        .join(OrderLine, PlanEntry.order_line_id == OrderLine.id)
        .filter(PlanEntry.period == period)
        .options(
            selectinload(ExecutionEntry.plan_entry)
            .selectinload(PlanEntry.order_line)
            .selectinload(OrderLine.contract),
            selectinload(ExecutionEntry.plan_entry)
            .selectinload(PlanEntry.order_line)
            .selectinload(OrderLine.work_type),
        )
        .order_by(OrderLine.contract_id, OrderLine.id)
    )
    if contract_id:
        qset = qset.filter(OrderLine.contract_id == contract_id)
    entries = qset.all()

    # Группировка по договору

    groups: dict[int, list[ExecutionEntry]] = OrderedDict()
    for ex in entries:
        cid = ex.plan_entry.order_line.contract_id
        groups.setdefault(cid, []).append(ex)

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт о выполнении"
    N = 10

    for i, w in enumerate([4, 38, 20, 7, 10, 13, 10, 13, 7, 12], 1):
        _col_w(ws, i, w)

    # ── Заголовок ────────────────────────────────────────────────────
    _row_h(ws, 1, 4)
    _row_h(ws, 2, 28)
    _merge(ws, 2, 1, 2, N, "ОПЕРАТИВНЫЙ ОТЧЁТ О ВЫПОЛНЕНИИ РАБОТ", bold=True, size=14)
    _row_h(ws, 3, 16)
    _merge(ws, 3, 1, 3, N, f"за {_period_ru(period, genitive=True).lower()}", size=11)
    _row_h(ws, 4, 6)

    # ── Шапка таблицы ────────────────────────────────────────────────
    headers = [
        "№", "Наименование работы", "Объект", "Ед.\nизм.",
        "Кол-во\nплан", "Сумма\nплан, руб.",
        "Кол-во\nфакт", "Сумма\nфакт, руб.",
        "%\nвып.", "Статус",
    ]
    r = 5
    _row_h(ws, r, 28)
    for i, h in enumerate(headers, 1):
        _cell(ws, r, i, h, bold=True, size=9, h_align="center", wrap=True, fill=_FILL_HEADER)

    # ── Данные ───────────────────────────────────────────────────────
    r = 6
    grand_plan_qty = grand_plan_sum = grand_fact_qty = grand_fact_sum = 0.0
    n_global = 0

    for cid, exs in groups.items():
        c = exs[0].plan_entry.order_line.contract

        _row_h(ws, r, 15)
        label = f"{c.counterparty or '—'}   {c.contract_no or c.number}"
        _merge(ws, r, 1, r, N, label, bold=True, size=9,
               fill=_FILL_GROUP, h_align="left", border=True)
        _table_borders(ws, r, r, 1, N)
        r += 1

        g_pq = g_ps = g_fq = g_fs = 0.0

        for ex in exs:
            n_global += 1
            pe = ex.plan_entry
            ol = pe.order_line
            u = unit_price(ol)
            plan_qty = float(pe.plan_qty or 0)
            plan_sum = (pe.plan_sum or 0) or plan_qty * u
            fact_qty = float(ex.qty_fact or 0)
            fact_sum = float(ex.sum_fact or 0)
            pct = round(fact_sum / plan_sum * 100, 1) if plan_sum else 0.0
            signed = ex.status == ExecutionStatus.SIGNED
            row_fill = _FILL_SIGNED if signed else None

            g_pq += plan_qty; g_ps += plan_sum
            g_fq += fact_qty; g_fs += fact_sum

            _cell(ws, r, 1, n_global, h_align="center", fill=row_fill)
            _cell(ws, r, 2, ol.description, h_align="left", wrap=True, fill=row_fill)
            _cell(ws, r, 3, c.project or "", h_align="left", wrap=True, fill=row_fill)
            _cell(ws, r, 4, ol.work_type.unit if ol.work_type and ol.work_type.unit else "—",
                  h_align="center", fill=row_fill)
            _cell(ws, r, 5, round(plan_qty, 2), h_align="right", fill=row_fill, num_fmt="#,##0.00")
            _cell(ws, r, 6, round(plan_sum, 2), h_align="right", fill=row_fill, num_fmt="#,##0.00")
            _cell(ws, r, 7, round(fact_qty, 2) if fact_qty else None,
                  h_align="right", fill=row_fill, num_fmt="#,##0.00")
            _cell(ws, r, 8, round(fact_sum, 2) if fact_sum else None,
                  h_align="right", fill=row_fill, num_fmt="#,##0.00")
            _cell(ws, r, 9, pct if pct else None, h_align="right", fill=row_fill, num_fmt="0.0")
            status_text = "подписан" if signed else ("черновик" if fact_qty else "—")
            _cell(ws, r, 10, status_text, h_align="center", fill=row_fill)
            r += 1

        # Итого по договору
        g_pct = round(g_fs / g_ps * 100, 1) if g_ps else 0.0
        _row_h(ws, r, 14)
        _merge(ws, r, 1, r, 4, f"Итого: {c.counterparty or '—'}", bold=True, size=9,
               fill=_FILL_TOTAL, h_align="left", border=True)
        _table_borders(ws, r, r, 1, 4)
        _cell(ws, r, 5, round(g_pq, 2), bold=True, fill=_FILL_TOTAL, h_align="right", num_fmt="#,##0.00")
        _cell(ws, r, 6, round(g_ps, 2), bold=True, fill=_FILL_TOTAL, h_align="right", num_fmt="#,##0.00")
        _cell(ws, r, 7, round(g_fq, 2), bold=True, fill=_FILL_TOTAL, h_align="right", num_fmt="#,##0.00")
        _cell(ws, r, 8, round(g_fs, 2), bold=True, fill=_FILL_TOTAL, h_align="right", num_fmt="#,##0.00")
        _cell(ws, r, 9, g_pct, bold=True, fill=_FILL_TOTAL, h_align="right", num_fmt="0.0")
        _cell(ws, r, 10, None, fill=_FILL_TOTAL)

        grand_plan_qty += g_pq; grand_plan_sum += g_ps
        grand_fact_qty += g_fq; grand_fact_sum += g_fs
        r += 1

    # ── ИТОГО общий ──────────────────────────────────────────────────
    g_pct = round(grand_fact_sum / grand_plan_sum * 100, 1) if grand_plan_sum else 0.0
    _row_h(ws, r, 16)
    _merge(ws, r, 1, r, 4, "ИТОГО ПО ВСЕМ ДОГОВОРАМ", bold=True, size=10,
           fill=_FILL_HEADER, h_align="left", border=True)
    _table_borders(ws, r, r, 1, 4)
    _cell(ws, r, 5, round(grand_plan_qty, 2), bold=True, size=10, fill=_FILL_HEADER, h_align="right", num_fmt="#,##0.00")
    _cell(ws, r, 6, round(grand_plan_sum, 2), bold=True, size=10, fill=_FILL_HEADER, h_align="right", num_fmt="#,##0.00")
    _cell(ws, r, 7, round(grand_fact_qty, 2), bold=True, size=10, fill=_FILL_HEADER, h_align="right", num_fmt="#,##0.00")
    _cell(ws, r, 8, round(grand_fact_sum, 2), bold=True, size=10, fill=_FILL_HEADER, h_align="right", num_fmt="#,##0.00")
    _cell(ws, r, 9, g_pct, bold=True, size=10, fill=_FILL_HEADER, h_align="right", num_fmt="0.0")
    _cell(ws, r, 10, None, fill=_FILL_HEADER)

    _a4(ws, landscape=True)
    ws.print_title_rows = "5:5"
    ws.freeze_panes = "A6"

    return _finish(wb)
