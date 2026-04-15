from __future__ import annotations

from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_
from sqlalchemy.orm import Session, selectinload

from app.models import ActLine, Contract, ExecutionEntry, OrderLine, PlanEntry, WorkType
from app.services.balances import balance
from app.services.pricing import unit_price


HEADER_FILL = PatternFill("solid", fgColor="E2E8F0")
HEADER_FONT = Font(bold=True)


def _autofit(ws, max_width: int = 60) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 10
        for cell in col:
            v = cell.value
            if v is None:
                continue
            length = max(length, min(max_width, len(str(v)) + 2))
        ws.column_dimensions[letter].width = length


def _header(ws, row: int, titles: list[str]) -> None:
    for i, t in enumerate(titles, start=1):
        c = ws.cell(row=row, column=i, value=t)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")


def _finish(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_contracts(session: Session, contract_id: Optional[int] = None, q: Optional[str] = None) -> bytes:
    cset = session.query(Contract)
    if contract_id:
        cset = cset.filter(Contract.id == contract_id)
    if q:
        like = f"%{q}%"
        cset = cset.filter(or_(
            Contract.number.ilike(like),
            Contract.contract_no.ilike(like),
            Contract.counterparty.ilike(like),
            Contract.project.ilike(like),
            Contract.order_label.ilike(like),
        ))
    contracts = cset.options(
        selectinload(Contract.order_lines).selectinload(OrderLine.work_type)
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Договоры"
    _header(ws, 1, [
        "Договор №", "Заказ", "Контрагент", "Проект", "Статус",
        "Работа", "Тип работы",
        "Кол-во заказано", "Сумма заказано, ₽",
        "Кол-во выполнено", "Сумма выполнено, ₽",
        "Кол-во остаток", "Сумма остаток, ₽",
        "План, ₽", "Подписано, ₽",
    ])
    r = 2
    for c in contracts:
        for ol in c.order_lines:
            b = balance(session, ol.id)
            unit = unit_price(ol)
            pe_list = session.query(PlanEntry).filter(PlanEntry.order_line_id == ol.id).all()
            plan_rub = sum(float(pe.plan_qty or 0.0) * unit for pe in pe_list)
            pe_ids = [pe.id for pe in pe_list]
            signed_rub = 0.0
            if pe_ids:
                signed_qty = session.query(func.coalesce(func.sum(ExecutionEntry.qty_fact), 0.0)).filter(
                    ExecutionEntry.plan_entry_id.in_(pe_ids),
                    ExecutionEntry.status == "signed",
                    ExecutionEntry.act_line_id.is_(None),
                ).scalar() or 0.0
                signed_rub = float(signed_qty) * unit
            ws.append([
                c.contract_no or c.number, c.order_label, c.counterparty, c.project, c.status,
                ol.description, (ol.work_type.name if ol.work_type else None),
                float(ol.qty_ordered or 0.0), float(ol.sum_ordered or 0.0),
                float(b.qty_done or 0.0), float(b.sum_done or 0.0),
                float(b.qty_remaining or 0.0), float(b.sum_remaining or 0.0),
                plan_rub, signed_rub,
            ])
            r += 1
    _autofit(ws)
    ws.freeze_panes = "A2"
    return _finish(wb)


def export_plan(session: Session, periods: list[str], contract_id: Optional[int] = None, q: Optional[str] = None) -> bytes:
    qset = (
        session.query(OrderLine)
        .options(selectinload(OrderLine.contract), selectinload(OrderLine.work_type), selectinload(OrderLine.plan_entries))
    )
    if contract_id:
        qset = qset.filter(OrderLine.contract_id == contract_id)
    if q:
        like = f"%{q}%"
        qset = qset.join(Contract).join(WorkType).filter(or_(
            OrderLine.description.ilike(like),
            WorkType.name.ilike(like),
            Contract.number.ilike(like),
            Contract.counterparty.ilike(like),
            Contract.project.ilike(like),
        ))
    order_lines = qset.order_by(OrderLine.contract_id, OrderLine.id).all()

    plans_by_key = {(p.order_line_id, p.period): p for ol in order_lines for p in ol.plan_entries}
    plan_ids = [p.id for p in plans_by_key.values()]
    ex_by_plan: dict[int, ExecutionEntry] = {}
    if plan_ids:
        ex_by_plan = {
            ex.plan_entry_id: ex
            for ex in session.query(ExecutionEntry).filter(ExecutionEntry.plan_entry_id.in_(plan_ids)).all()
        }

    wb = Workbook()
    ws = wb.active
    ws.title = "План"
    head = ["Договор", "Контрагент", "Проект", "Работа", "Тип работы",
            "Кол-во заказано", "Кол-во остаток", "Сумма остаток, ₽"]
    for p in periods:
        head += [f"План {p}", f"Статус {p}", f"Факт {p}", f"Подписан {p}"]
    _header(ws, 1, head)

    for ol in order_lines:
        b = balance(session, ol.id)
        row = [
            (ol.contract.number if ol.contract else None),
            (ol.contract.counterparty if ol.contract else None),
            (ol.contract.project if ol.contract else None),
            ol.description,
            (ol.work_type.name if ol.work_type else None),
            float(ol.qty_ordered or 0.0),
            float(b.qty_remaining or 0.0),
            float(b.sum_remaining or 0.0),
        ]
        for p in periods:
            pe = plans_by_key.get((ol.id, p))
            ex = ex_by_plan.get(pe.id) if pe else None
            row += [
                (float(pe.plan_qty) if pe and pe.plan_qty is not None else None),
                (pe.status if pe else None),
                (float(ex.qty_fact) if ex and ex.qty_fact is not None else None),
                ("да" if ex and ex.status == "signed" else ""),
            ]
        ws.append(row)
    _autofit(ws)
    ws.freeze_panes = "A2"
    return _finish(wb)


def export_execution(session: Session, period: str, contract_id: Optional[int] = None) -> bytes:
    qset = (
        session.query(ExecutionEntry)
        .join(PlanEntry, ExecutionEntry.plan_entry_id == PlanEntry.id)
        .join(OrderLine, PlanEntry.order_line_id == OrderLine.id)
        .filter(PlanEntry.period == period)
        .options(
            selectinload(ExecutionEntry.plan_entry)
            .selectinload(PlanEntry.order_line)
            .selectinload(OrderLine.contract)
        )
    )
    if contract_id:
        qset = qset.filter(OrderLine.contract_id == contract_id)
    entries = qset.order_by(OrderLine.contract_id, OrderLine.id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Исполнение {period}"
    _header(ws, 1, [
        "Договор", "Контрагент", "Проект", "Работа",
        "План кол-во", "План сумма, ₽",
        "Факт кол-во", "Факт сумма, ₽",
        "% выполнения", "Статус", "Комментарий",
    ])
    for ex in entries:
        pe = ex.plan_entry
        ol = pe.order_line
        c = ol.contract
        unit = unit_price(ol)
        plan_qty = pe.plan_qty or 0.0
        plan_sum = (pe.plan_sum or 0.0) or plan_qty * unit
        fact_qty = ex.qty_fact or 0.0
        fact_sum = ex.sum_fact or 0.0
        pct = (fact_sum / plan_sum * 100) if plan_sum else 0.0
        ws.append([
            (c.number if c else None),
            (c.counterparty if c else None),
            (c.project if c else None),
            ol.description,
            float(plan_qty), float(plan_sum),
            float(fact_qty), float(fact_sum),
            round(pct, 1),
            ("подписан" if ex.status == "signed" else "черновик"),
            ex.note or "",
        ])
    _autofit(ws)
    ws.freeze_panes = "A2"
    return _finish(wb)


def export_dashboard(session: Session, periods_filter: Optional[list[str]]) -> bytes:
    qry = (
        session.query(PlanEntry, OrderLine, Contract, ExecutionEntry)
        .join(OrderLine, PlanEntry.order_line_id == OrderLine.id)
        .join(Contract, OrderLine.contract_id == Contract.id)
        .outerjoin(ExecutionEntry, ExecutionEntry.plan_entry_id == PlanEntry.id)
    )
    if periods_filter is not None:
        qry = qry.filter(PlanEntry.period.in_(periods_filter))
    rows = qry.all()

    by_period: dict[str, dict] = {}
    by_contract: dict[int, dict] = {}
    for pe, ol, c, ex in rows:
        unit = unit_price(ol)
        plan_qty = pe.plan_qty or 0.0
        plan_sum = (pe.plan_sum or 0.0) or plan_qty * unit
        fact_sum = (ex.sum_fact or 0.0) if ex else 0.0
        signed = bool(ex and ex.status == "signed")
        approved = pe.status == "approved"

        p = by_period.setdefault(pe.period, {"period": pe.period, "plan_draft": 0.0, "plan_approved": 0.0, "fact": 0.0, "signed": 0.0})
        k = by_contract.setdefault(c.id, {"contract": c, "plan": 0.0, "fact": 0.0, "signed": 0.0})
        if approved:
            p["plan_approved"] += plan_sum
        else:
            p["plan_draft"] += plan_sum
        k["plan"] += plan_sum
        p["fact"] += fact_sum
        k["fact"] += fact_sum
        if signed:
            p["signed"] += fact_sum
            k["signed"] += fact_sum

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "По периодам"
    _header(ws1, 1, ["Период", "План черновик, ₽", "План утверждён, ₽", "План итого, ₽", "Факт, ₽", "Подписано, ₽", "% факт", "% подписано"])
    for p in sorted(by_period.values(), key=lambda x: x["period"]):
        total_plan = p["plan_draft"] + p["plan_approved"]
        ws1.append([
            p["period"], p["plan_draft"], p["plan_approved"], total_plan,
            p["fact"], p["signed"],
            round(p["fact"] / total_plan * 100, 1) if total_plan else 0,
            round(p["signed"] / total_plan * 100, 1) if total_plan else 0,
        ])
    _autofit(ws1)
    ws1.freeze_panes = "A2"

    ws2 = wb.create_sheet("По договорам")
    _header(ws2, 1, ["Договор №", "Контрагент", "Проект", "План, ₽", "Факт, ₽", "Подписано, ₽", "% подписано"])
    for k in sorted(by_contract.values(), key=lambda x: x["plan"], reverse=True):
        c = k["contract"]
        ws2.append([
            c.number, c.counterparty, c.project,
            k["plan"], k["fact"], k["signed"],
            round(k["signed"] / k["plan"] * 100, 1) if k["plan"] else 0,
        ])
    _autofit(ws2)
    ws2.freeze_panes = "A2"
    return _finish(wb)
